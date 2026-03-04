#!/usr/bin/env python3
"""
Sprint Summary — Generate an Excel report with task breakdown and time analysis.

Usage:
    python3 sprint_summary.py --sprint-number 45 --start-date 2026-02-24 --end-date 2026-03-07

Requires: TEAMWORK_SITE and TEAMWORK_API_KEY environment variables.
Requires: openpyxl (pip install openpyxl)

Optimized to use the Teamwork workflows endpoint with sideloaded data
(timeTotals, projects, tasklists, tags, users) to minimize API calls.
Typically completes in ~12-26 API calls instead of ~100-170.
"""

import argparse
import json
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import tw_api

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BOARD_STATUS_COMPLETE = ("On Production", "Done")
BOARD_STATUS_STAGING = "On Staging"
BOARD_STATUS_READY_PROD = "Ready for Production"

TIME_SUMMARY_PEOPLE = [
    "Rodolfo Ortiz",
    "Ulises Becerra",
    "Fernando Mendez",
]


# ---------------------------------------------------------------------------
# Tab 1 helpers
# ---------------------------------------------------------------------------

def find_current_and_previous_tags(sprint_number):
    """
    Find the Teamwork tags for the current and previous sprint.

    Sprint tags follow the pattern "Sprint {N} - {date}" or "Sprint {N}".
    We match any tag containing "Sprint {N}" (case-insensitive).

    Returns:
        (current_tag, previous_tag) — each is a tag dict or None
    """
    current_pattern = f"Sprint {sprint_number}"
    previous_pattern = f"Sprint {sprint_number - 1}"

    current_matches = tw_api.find_sprint_tags(current_pattern)
    previous_matches = tw_api.find_sprint_tags(previous_pattern)

    current_tag = None
    previous_tag = None

    # Prefer exact match on the number portion
    for t in current_matches:
        name = t.get("name", "")
        if current_pattern.lower() in name.lower():
            current_tag = t
            break

    for t in previous_matches:
        name = t.get("name", "")
        if previous_pattern.lower() in name.lower():
            previous_tag = t
            break

    return current_tag, previous_tag


def categorize_tasks(tasks, current_tag_id, previous_tag_id, tag_map=None):
    """
    Categorize sprint tasks into Carryover, Planned, and Unplanned.

    Supports both regular task format (tags as objects with name/id) and
    workflows format (tagIds as flat int array + sideloaded tag_map).

    Args:
        tasks: list of task dicts
        current_tag_id: ID of the current sprint tag
        previous_tag_id: ID of the previous sprint tag (may be None)
        tag_map: optional {tag_id: {id, name, ...}} from workflows sideloaded data

    Returns:
        dict with keys 'Carryover', 'Planned', 'Unplanned' mapping to lists
    """
    categories = {"Carryover": [], "Planned": [], "Unplanned": []}

    for task in tasks:
        # Get tag IDs and names — handle both response formats
        task_tag_ids = task.get("tagIds") or []
        tag_objects = task.get("tags") or []

        if task_tag_ids:
            # Workflows format (or any format with tagIds): flat [int] array
            tag_ids = [int(tid) for tid in task_tag_ids]
            if tag_map:
                tag_names = []
                for tid in tag_ids:
                    tag_info = tag_map.get(int(tid), {})
                    name = tag_info.get("name", "") if isinstance(tag_info, dict) else ""
                    tag_names.append(name.lower())
            else:
                # No tag_map available — IDs still work for carryover detection
                tag_names = []
        elif tag_objects:
            # Regular format: tags are [{id, name, ...}] objects
            tag_names = [t.get("name", "").lower() for t in tag_objects if isinstance(t, dict)]
            tag_ids = [int(t.get("id", 0)) for t in tag_objects if isinstance(t, dict)]
        else:
            tag_names = []
            tag_ids = []

        has_unplanned = any("unplanned" in n for n in tag_names)
        has_previous = (previous_tag_id is not None and
                        int(previous_tag_id) in tag_ids)

        if has_unplanned:
            categories["Unplanned"].append(task)
        elif has_previous:
            categories["Carryover"].append(task)
        else:
            categories["Planned"].append(task)

    return categories


def classify_task_status(task, board_status_map):
    """
    Classify a single task as Complete, Incomplete, On Staging, or Ready for Production.

    Args:
        task: task dict
        board_status_map: dict {task_id: column_name/stage_name}

    Returns:
        One of: 'Complete', 'Incomplete', 'On Staging', 'Ready for Production'
    """
    task_id = int(task.get("id", 0))
    status = task.get("status", "").lower()
    board_col = board_status_map.get(task_id, "")

    # Complete: task status is completed, or board column is On Production / Done
    if status == "completed" or board_col in BOARD_STATUS_COMPLETE:
        return "Complete"

    # On Staging: open task in the On Staging column
    if board_col == BOARD_STATUS_STAGING:
        return "On Staging"

    # Ready for Production: open task in the Ready for Production column
    if board_col == BOARD_STATUS_READY_PROD:
        return "Ready for Production"

    # Everything else is incomplete
    return "Incomplete"


def build_task_summary_rows(sprint_number, categories, board_status_map, time_totals_map):
    """
    Build the data rows for Tab 1.

    Uses timeTotals from the workflows endpoint for logged hours (includes
    ALL users' time on each task, not filtered).

    Args:
        sprint_number: int
        categories: dict from categorize_tasks()
        board_status_map: {task_id: stage_name}
        time_totals_map: {task_id: {loggedMinutes: N, ...}} from workflows sideloaded data

    Returns:
        list of dicts, one per task type, plus a totals row
    """
    rows = []

    for task_type in ("Carryover", "Planned", "Unplanned"):
        tasks = categories[task_type]
        total = len(tasks)

        completed = []
        incomplete = []
        on_staging = []
        ready_prod = []

        for task in tasks:
            classification = classify_task_status(task, board_status_map)
            if classification == "Complete":
                completed.append(task)
            elif classification == "On Staging":
                on_staging.append(task)
            elif classification == "Ready for Production":
                ready_prod.append(task)
            else:
                incomplete.append(task)

        # Time for completed tasks
        estimated_mins = 0
        logged_mins = 0
        for task in completed:
            # Handle both field names: estimateMinutes (workflows) vs estimatedMinutes (regular)
            estimated_mins += (task.get("estimateMinutes")
                               or task.get("estimatedMinutes")
                               or 0)
            task_id = int(task.get("id", 0))
            # Use timeTotals for logged time (includes ALL users)
            totals = time_totals_map.get(task_id, {})
            logged_mins += totals.get("loggedMinutes", 0) or 0

        estimated_hours = round(estimated_mins / 60, 2) if estimated_mins else 0
        logged_hours = round(logged_mins / 60, 2) if logged_mins else 0
        diff_hours = round(logged_hours - estimated_hours, 2)
        logged_vs_estimated_pct = round(logged_hours / estimated_hours, 4) if estimated_hours > 0 else 0

        rows.append({
            "sprint_number": sprint_number,
            "task_type": task_type,
            "total": total,
            "completed_num": len(completed),
            "completed_pct": len(completed) / total if total > 0 else 0,
            "incomplete_num": len(incomplete),
            "incomplete_pct": len(incomplete) / total if total > 0 else 0,
            "on_staging_num": len(on_staging),
            "on_staging_pct": len(on_staging) / total if total > 0 else 0,
            "ready_prod_num": len(ready_prod),
            "ready_prod_pct": len(ready_prod) / total if total > 0 else 0,
            "estimated_hours": estimated_hours,
            "logged_hours": logged_hours,
            "logged_vs_estimated": diff_hours,
            "logged_vs_estimated_pct": logged_vs_estimated_pct,
        })

    # Totals row
    t_total = sum(r["total"] for r in rows)
    t_completed = sum(r["completed_num"] for r in rows)
    t_incomplete = sum(r["incomplete_num"] for r in rows)
    t_staging = sum(r["on_staging_num"] for r in rows)
    t_ready = sum(r["ready_prod_num"] for r in rows)
    t_est = sum(r["estimated_hours"] for r in rows)
    t_log = sum(r["logged_hours"] for r in rows)
    t_diff = round(t_log - t_est, 2)
    t_pct = round(t_log / t_est, 4) if t_est > 0 else 0

    rows.append({
        "sprint_number": sprint_number,
        "task_type": "TOTAL",
        "total": t_total,
        "completed_num": t_completed,
        "completed_pct": t_completed / t_total if t_total > 0 else 0,
        "incomplete_num": t_incomplete,
        "incomplete_pct": t_incomplete / t_total if t_total > 0 else 0,
        "on_staging_num": t_staging,
        "on_staging_pct": t_staging / t_total if t_total > 0 else 0,
        "ready_prod_num": t_ready,
        "ready_prod_pct": t_ready / t_total if t_total > 0 else 0,
        "estimated_hours": round(t_est, 2),
        "logged_hours": round(t_log, 2),
        "logged_vs_estimated": t_diff,
        "logged_vs_estimated_pct": t_pct,
    })

    return rows


# ---------------------------------------------------------------------------
# Tab 2 helpers
# ---------------------------------------------------------------------------

def find_person_ids_from_user_map(names, user_map):
    """
    Look up Teamwork person IDs by full name using the sideloaded user map.

    Args:
        names: list of full name strings
        user_map: {user_id: {id, firstName, lastName, ...}} from workflows data

    Returns:
        dict: {full_name: person_id} (None if not found)
    """
    result = {}
    for name in names:
        parts = name.lower().split()
        for uid, user in user_map.items():
            first = (user.get("firstName") or "").lower()
            last = (user.get("lastName") or "").lower()
            if len(parts) == 2 and first == parts[0] and last == parts[1]:
                result[name] = int(uid)
                break
        if name not in result:
            print(f"WARNING: Could not find person '{name}' in sideloaded users.", file=sys.stderr)
            result[name] = None
    return result


def is_non_billable(task_name, tasklist_name, project_name):
    """
    Determine if a time entry is non-billable based on naming conventions.

    Non-billable if any of:
    - task name contains 'Non-Billable'
    - task list name contains 'Non-Billable'
    - project name contains 'Non-Billable'
    - project name starts with 'URI-'
    """
    task_name = (task_name or "").strip()
    tasklist_name = (tasklist_name or "").strip()
    project_name = (project_name or "").strip()

    if "non-billable" in task_name.lower():
        return True
    if "non-billable" in tasklist_name.lower():
        return True
    if "non-billable" in project_name.lower():
        return True
    if project_name.upper().startswith("URI-"):
        return True
    return False


def build_time_summary_rows(sprint_number, current_tag_id, person_ids,
                            time_entries_by_person, task_map,
                            project_name_map, tasklist_name_map, tag_name_map):
    """
    Build the data rows for Tab 2 using pre-fetched data (ZERO API calls).

    Args:
        sprint_number: int
        current_tag_id: ID of the current sprint tag
        person_ids: {full_name: person_id}
        time_entries_by_person: {user_id: [time_entry_dicts]}
        task_map: {task_id: task_dict} (sprint + non-sprint tasks)
        project_name_map: {project_id: project_name}
        tasklist_name_map: {tasklist_id: tasklist_name}
        tag_name_map: {tag_id: tag_name}

    Returns:
        list of dicts, one per person
    """
    rows = []

    for person_name in TIME_SUMMARY_PEOPLE:
        person_id = person_ids.get(person_name)
        if person_id is None:
            rows.append({
                "sprint_number": sprint_number,
                "person": person_name,
                "total_hours": 0,
                "billable_hours": 0,
                "non_billable_hours": 0,
                "planned_hours": 0,
                "unplanned_hours": 0,
                "other_hours": 0,
            })
            continue

        entries = time_entries_by_person.get(int(person_id), [])

        total_mins = 0
        billable_mins = 0
        non_billable_mins = 0
        planned_mins = 0
        unplanned_mins = 0

        for entry in entries:
            mins = entry.get("minutes", 0) or 0
            total_mins += mins

            task_id = entry.get("taskId") or entry.get("task-id") or entry.get("todoItemId")
            project_id = entry.get("projectId") or entry.get("project-id")

            task_name = ""
            tasklist_name = ""
            project_name = ""
            task_tag_ids = []

            # Look up task details from pre-fetched map (no API call)
            if task_id:
                task_id = int(task_id)
                task_detail = task_map.get(task_id)
                if task_detail:
                    task_name = task_detail.get("name", "")
                    # Get tag IDs — handle both formats
                    task_tag_ids = task_detail.get("tagIds") or []
                    if not task_tag_ids:
                        # Fall back to extracting IDs from tag objects
                        tag_objs = task_detail.get("tags") or []
                        task_tag_ids = [int(t.get("id", 0)) for t in tag_objs
                                        if isinstance(t, dict) and t.get("id")]
                    # Get tasklist name from pre-fetched map
                    tl_id = (task_detail.get("tasklistId")
                             or task_detail.get("taskListId")
                             or task_detail.get("todoListId"))
                    if tl_id:
                        tasklist_name = tasklist_name_map.get(int(tl_id), "")
                    if not project_id:
                        project_id = task_detail.get("projectId")

            # Look up project name from pre-fetched map (no API call)
            if project_id:
                project_name = project_name_map.get(int(project_id), "")

            # Classify billable vs non-billable
            if is_non_billable(task_name, tasklist_name, project_name):
                non_billable_mins += mins
            else:
                billable_mins += mins

            # Classify planned vs unplanned vs other
            tag_ids = [int(tid) for tid in task_tag_ids]
            tag_names = []
            for tid in tag_ids:
                name = tag_name_map.get(int(tid), "")
                tag_names.append(name.lower())

            has_current_sprint = int(current_tag_id) in tag_ids if current_tag_id else False
            has_unplanned = any("unplanned" in n for n in tag_names)

            if has_unplanned:
                unplanned_mins += mins
            elif has_current_sprint:
                planned_mins += mins
            # else: other (calculated as remainder)

        other_mins = total_mins - planned_mins - unplanned_mins

        rows.append({
            "sprint_number": sprint_number,
            "person": person_name,
            "total_hours": round(total_mins / 60, 2),
            "billable_hours": round(billable_mins / 60, 2),
            "non_billable_hours": round(non_billable_mins / 60, 2),
            "planned_hours": round(planned_mins / 60, 2),
            "unplanned_hours": round(unplanned_mins / 60, 2),
            "other_hours": round(other_mins / 60, 2),
        })

    return rows


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def create_workbook(task_rows, time_rows, sprint_number):
    """
    Create the Excel workbook with both tabs.

    Returns:
        Workbook object
    """
    wb = Workbook()

    # -- Styles --
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    pct_format = '0.0%'
    hours_format = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # =====================================================================
    # Tab 1 — Sprint Task Summary
    # =====================================================================
    ws1 = wb.active
    ws1.title = "Sprint Task Summary"

    headers_tab1 = [
        "Sprint #",
        "Task Type",
        "Total Tasks",
        "Completed #",
        "Completed %",
        "Incomplete #",
        "Incomplete %",
        "On Staging #",
        "On Staging %",
        "Ready for Production #",
        "Ready for Production %",
        "Completed Tasks Estimated Hours",
        "Completed Tasks Logged Hours",
        "Completed Tasks Logged vs. Estimated",
        "Logged vs. Estimated %",
    ]

    for col_idx, header in enumerate(headers_tab1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, row_data in enumerate(task_rows, 2):
        is_total = row_data["task_type"] == "TOTAL"
        bold = Font(bold=True, size=11) if is_total else Font(size=11)

        values = [
            row_data["sprint_number"],
            row_data["task_type"],
            row_data["total"],
            row_data["completed_num"],
            row_data["completed_pct"],
            row_data["incomplete_num"],
            row_data["incomplete_pct"],
            row_data["on_staging_num"],
            row_data["on_staging_pct"],
            row_data["ready_prod_num"],
            row_data["ready_prod_pct"],
            row_data["estimated_hours"],
            row_data["logged_hours"],
            row_data["logged_vs_estimated"],
            row_data["logged_vs_estimated_pct"],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Format percentage columns
            if col_idx in (5, 7, 9, 11, 15):
                cell.number_format = pct_format
            # Format hours columns
            elif col_idx in (12, 13, 14):
                cell.number_format = hours_format

    # Auto-size columns
    for col_idx in range(1, len(headers_tab1) + 1):
        max_len = len(str(headers_tab1[col_idx - 1]))
        for row_idx in range(2, len(task_rows) + 2):
            val = ws1.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws1.column_dimensions[ws1.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    # =====================================================================
    # Tab 2 — Sprint Time Summary
    # =====================================================================
    ws2 = wb.create_sheet(title="Sprint Time Summary")

    headers_tab2 = [
        "Sprint Number",
        "Person",
        "Total Hours",
        "Total Billable Hours",
        "Total Non-Billable Hours",
        "Planned Hours",
        "Unplanned Hours",
        "Other",
    ]

    for col_idx, header in enumerate(headers_tab2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, row_data in enumerate(time_rows, 2):
        values = [
            row_data["sprint_number"],
            row_data["person"],
            row_data["total_hours"],
            row_data["billable_hours"],
            row_data["non_billable_hours"],
            row_data["planned_hours"],
            row_data["unplanned_hours"],
            row_data["other_hours"],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=11)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Format hours columns
            if col_idx >= 3:
                cell.number_format = hours_format

    # Auto-size columns
    for col_idx in range(1, len(headers_tab2) + 1):
        max_len = len(str(headers_tab2[col_idx - 1]))
        for row_idx in range(2, len(time_rows) + 2):
            val = ws2.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def sprint_summary(sprint_number, start_date, end_date):
    """
    Generate the full sprint summary Excel report.

    Optimized flow:
      1. Find sprint tags
      2. Fetch all workflow tasks with sideloaded data (timeTotals, projects,
         tasklists, tags, users) — replaces separate task, board, project,
         tasklist, and people lookups.
      3. Resolve person IDs from sideloaded users
      4. Fetch time entries for team members (1 call with comma-separated userIds)
      5. Batch fetch non-sprint task details (only tasks the 3 users logged time on)
      6. Process everything in memory, generate Excel report
    """
    print(f"Generating Sprint {sprint_number} summary ({start_date} to {end_date})...", file=sys.stderr)

    # ==================================================================
    # Phase 1: Data Collection
    # ==================================================================

    # ------------------------------------------------------------------
    # Step 1: Find sprint tags (1 API call)
    # ------------------------------------------------------------------
    print("  [1/6] Finding sprint tags...", file=sys.stderr)
    current_tag, previous_tag = find_current_and_previous_tags(sprint_number)

    if current_tag is None:
        print(f"ERROR: Could not find a tag matching 'Sprint {sprint_number}'.", file=sys.stderr)
        print("Available sprint tags:", file=sys.stderr)
        for t in tw_api.find_sprint_tags():
            print(f"  - {t.get('name')} (id: {t.get('id')})", file=sys.stderr)
        sys.exit(1)

    current_tag_id = current_tag.get("id")
    previous_tag_id = previous_tag.get("id") if previous_tag else None

    print(f"  Current sprint tag: {current_tag.get('name')} (id: {current_tag_id})", file=sys.stderr)
    if previous_tag:
        print(f"  Previous sprint tag: {previous_tag.get('name')} (id: {previous_tag_id})", file=sys.stderr)
    else:
        print(f"  WARNING: No previous sprint tag found for Sprint {sprint_number - 1}. "
              f"Carryover detection will be skipped.", file=sys.stderr)

    # ------------------------------------------------------------------
    # Step 2: Fetch all workflow tasks with sideloaded data (~5-7 API calls)
    # ------------------------------------------------------------------
    print("  [2/6] Fetching workflow tasks with sideloaded data...", file=sys.stderr)
    (tasks, time_totals_map, project_map, tasklist_map,
     tag_map, user_map, board_status_map) = tw_api.get_all_workflow_tasks(
        tag_id=current_tag_id
    )
    print(f"  Found {len(tasks)} tasks in sprint.", file=sys.stderr)

    # ------------------------------------------------------------------
    # Step 3: Resolve person IDs from sideloaded users (ZERO API calls)
    # ------------------------------------------------------------------
    print("  [3/6] Resolving person IDs from sideloaded users...", file=sys.stderr)
    person_ids = find_person_ids_from_user_map(TIME_SUMMARY_PEOPLE, user_map)

    # If any person wasn't found in sideloaded users, try a direct people API call
    missing_people = [name for name, pid in person_ids.items() if pid is None]
    if missing_people:
        print(f"  Falling back to people API for: {missing_people}", file=sys.stderr)
        people = tw_api.fetch_all_pages("/people.json", result_key="people")
        for name in missing_people:
            parts = name.lower().split()
            for person in people:
                first = (person.get("firstName") or "").lower()
                last = (person.get("lastName") or "").lower()
                if len(parts) == 2 and first == parts[0] and last == parts[1]:
                    person_ids[name] = int(person.get("id"))
                    break

    # ------------------------------------------------------------------
    # Step 4: Fetch time entries for team members (1 API call, user-filtered)
    # ------------------------------------------------------------------
    print("  [4/6] Fetching time entries for team members...", file=sys.stderr)
    valid_person_ids = [pid for pid in person_ids.values() if pid is not None]
    all_person_entries = tw_api.get_time_entries_by_date_range(
        start_date, end_date, user_ids=valid_person_ids
    )
    print(f"  Found {len(all_person_entries)} time entries for {len(valid_person_ids)} team members.", file=sys.stderr)

    # Group time entries by person_id
    time_entries_by_person = {}
    for entry in all_person_entries:
        uid = entry.get("userId") or entry.get("user-id")
        if uid:
            uid = int(uid)
            time_entries_by_person.setdefault(uid, []).append(entry)

    for name, pid in person_ids.items():
        if pid:
            count = len(time_entries_by_person.get(int(pid), []))
            print(f"    {name}: {count} time entries", file=sys.stderr)

    # ------------------------------------------------------------------
    # Step 5: Batch fetch non-sprint task details (small set, not 3,400)
    # ------------------------------------------------------------------
    # Build sprint task lookup
    sprint_task_map = {int(t["id"]): t for t in tasks}

    # Find tasks referenced in the 3 users' time entries that are NOT in the sprint set
    referenced_task_ids = set()
    for entries in time_entries_by_person.values():
        for entry in entries:
            tid = entry.get("taskId") or entry.get("task-id") or entry.get("todoItemId")
            if tid:
                referenced_task_ids.add(int(tid))

    missing_task_ids = referenced_task_ids - set(sprint_task_map.keys())
    if missing_task_ids:
        print(f"  [5/6] Fetching details for {len(missing_task_ids)} non-sprint tasks...", file=sys.stderr)
        extra_tasks = tw_api.get_tasks_batch(list(missing_task_ids))
        sprint_task_map.update(extra_tasks)

        # Extract project/tasklist info from extra tasks
        for task in extra_tasks.values():
            pid = task.get("projectId")
            if pid and int(pid) not in project_map:
                # We need the project name for billable classification
                proj = tw_api.get_project_by_id(pid)
                if proj:
                    project_map[int(pid)] = {"id": int(pid), "name": proj.get("name", "")}

            tl_id = task.get("tasklistId") or task.get("taskListId") or task.get("todoListId")
            if tl_id and int(tl_id) not in tasklist_map:
                tl = tw_api.get_tasklist_by_id(tl_id)
                if tl:
                    tasklist_map[int(tl_id)] = {"id": int(tl_id), "name": tl.get("name", "")}

            # Extract tag info from extra tasks
            tag_objs = task.get("tags") or []
            for tag_obj in tag_objs:
                if isinstance(tag_obj, dict) and tag_obj.get("id"):
                    tid = int(tag_obj["id"])
                    if tid not in tag_map:
                        tag_map[tid] = tag_obj
    else:
        print("  [5/6] No non-sprint tasks to fetch.", file=sys.stderr)

    # ==================================================================
    # Phase 2: Build Lookup Maps (ZERO API calls)
    # ==================================================================
    print("  [6/6] Building report...", file=sys.stderr)

    # Build project name map (flat: project_id -> name)
    project_name_map = {}
    for pid, proj in project_map.items():
        if isinstance(proj, dict):
            project_name_map[int(pid)] = proj.get("name", "")
        else:
            project_name_map[int(pid)] = str(proj)

    # Build tasklist name map (flat: tasklist_id -> name)
    tasklist_name_map = {}
    for tlid, tl in tasklist_map.items():
        if isinstance(tl, dict):
            tasklist_name_map[int(tlid)] = tl.get("name", "")
        else:
            tasklist_name_map[int(tlid)] = str(tl)

    # Build tag name map (flat: tag_id -> name)
    tag_name_map = {}
    for tagid, tag in tag_map.items():
        if isinstance(tag, dict):
            tag_name_map[int(tagid)] = tag.get("name", "")
        else:
            tag_name_map[int(tagid)] = str(tag)

    # ==================================================================
    # Phase 3: In-Memory Processing
    # ==================================================================

    # Categorize tasks
    categories = categorize_tasks(tasks, current_tag_id, previous_tag_id, tag_map=tag_map)
    for cat, cat_tasks in categories.items():
        print(f"    {cat}: {len(cat_tasks)} tasks", file=sys.stderr)

    # Build Tab 1 rows using timeTotals
    task_rows = build_task_summary_rows(sprint_number, categories, board_status_map, time_totals_map)

    # Build Tab 2 rows using pre-fetched data
    time_rows = build_time_summary_rows(
        sprint_number, current_tag_id, person_ids,
        time_entries_by_person, sprint_task_map,
        project_name_map, tasklist_name_map, tag_name_map
    )

    # ==================================================================
    # Phase 4: Report Generation
    # ==================================================================
    filename = f"Sprint_{sprint_number}_Summary.xlsx"
    print(f"  Writing Excel file: {filename}", file=sys.stderr)

    wb = create_workbook(task_rows, time_rows, sprint_number)
    wb.save(filename)

    # Print API call count for diagnostics
    print(f"  Total API calls: {tw_api.get_request_count()}", file=sys.stderr)
    print(f"  Done.", file=sys.stderr)

    # Output the filename to stdout for Claude to present to the user
    print(json.dumps({
        "status": "success",
        "file": filename,
        "sprint_number": sprint_number,
        "start_date": start_date,
        "end_date": end_date,
        "task_summary": task_rows,
        "time_summary": time_rows,
    }, indent=2))


def main():
    parser = argparse.ArgumentParser(description="Generate Sprint Summary Excel Report")
    parser.add_argument("--sprint-number", type=int, required=True, help="Sprint number (e.g., 45)")
    parser.add_argument("--start-date", required=True, help="Sprint start date (YYYY-MM-DD)")
    parser.add_argument("--end-date", required=True, help="Sprint end date (YYYY-MM-DD)")

    args = parser.parse_args()

    sprint_summary(args.sprint_number, args.start_date, args.end_date)


if __name__ == "__main__":
    main()
